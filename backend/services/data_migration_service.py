"""Non-destructive FinLedge data migrations run before the API becomes ready."""

from __future__ import annotations

import json
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .excel_utils import to_float as _amount
from .path_utils import get_data_dir

MIGRATION_VERSION = "v1.2.0"
SHEET_NAME = "Bank"
HEADERS = [
    "Date",
    "Category",
    "Amount",
    "Cumulative Amount",
    "Description",
    "Created Timestamp",
    "Last Updated Timestamp",
    "Updated Device",
]
SHARE_HEADERS = [
    "Date",
    "Share Name",
    "Category",
    "Per Unit Price",
    "ASBA Charge",
    "Allotted",
    "Buy/Sell",
    "Total Amount",
    "Profit/Loss",
    "Cumulative Profit",
    "Created Timestamp",
    "Last Updated Timestamp",
    "Updated Device",
]
PERSONAL_FINANCE_HEADERS = [
    "Date",
    "Flow Type",
    "Direction",
    "Category",
    "Amount",
    "Signed Amount",
    "Description",
    "Source",
    "Created Timestamp",
    "Last Updated Timestamp",
    "Source Ref",
    "Updated Device",
]
LEGACY_CATEGORIES = {"income", "service cost", "investment cost", "operation cost"}
RETIRED_CATEGORIES = {"atm charge", "sms charge"}


def _normalized(value: object) -> str:
    return str(value or "").strip().lower()


def _map_legacy_category(category: object, description: object) -> tuple[str, int, str]:
    original = _normalized(category)
    text = _normalized(description)
    compact_text = text.replace(" ", "")
    if original == "atm charge":
        return "Debit Card Charge", -1, "retired ATM Charge moved to Debit Card Charge"
    if original == "sms charge":
        return "Mobile Banking Charge", -1, "retired SMS Charge moved to Mobile Banking Charge"
    if original not in LEGACY_CATEGORIES:
        return str(category or "").strip(), 0, "already compatible"
    if original == "income" and ("interest" in text or "int" in text):
        return "Interest Earned", 1, "Income with interest/int"
    if original == "service cost" and "tax" in text:
        return "Interest Tax", -1, "Service Cost with tax"
    if original == "service cost" and "mobile banking" in text and "renew" in text:
        return "Mobile Banking Charge", -1, "Service Cost with mobile banking and renew"
    if original == "service cost" and "bank" in text and "renew" in text:
        return "Mobile Banking Charge", -1, "Service Cost with bank and renew"
    if original == "service cost" and "mobile" in text:
        return "Mobile Banking Charge", -1, "Service Cost with mobile"
    if original == "service cost" and "card" in text and "install" in text:
        return "Debit Card Charge", -1, "Service Cost with card and installment"
    if original == "investment cost" and "demat" in text and "meroshare" in compact_text and "renew" in text:
        return "Demat & MeroShare Renewal", -1, "Investment Cost with Demat, MeroShare, and renew"
    if original == "investment cost" and "meroshare" in compact_text and "renew" in text:
        return "MeroShare Renewal", -1, "Investment Cost with MeroShare and renew"
    if original == "investment cost" and "demat" in text and "renew" in text:
        return "Demat Renewal", -1, "Investment Cost with demat and renew"
    if original == "investment cost" and "broker" in text and "renew" in text:
        return "Broker Renewal", -1, "Investment Cost with broker and renew"
    return "Other Charges", -1, "legacy category did not match a specific v1.2.0 rule"


def _bank_needs_migration(source_path: Path) -> bool:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, max_col=len(HEADERS), values_only=True), ())
        headers = list(first_row)
        if headers != HEADERS:
            return True
        return any(
            _normalized(row[1] if len(row) > 1 else "") in LEGACY_CATEGORIES | RETIRED_CATEGORIES
            for row in sheet.iter_rows(min_row=2, values_only=True)
        )
    finally:
        workbook.close()


def _workbook_needs_header_migration(source_path: Path, sheet_name: str, headers: list[str]) -> bool:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, max_col=len(headers), values_only=True), ())
        return list(first_row) != headers
    finally:
        workbook.close()


def _read_and_map(source_path: Path) -> tuple[list[list[Any]], list[dict[str, Any]]]:
    workbook = load_workbook(source_path, data_only=True)
    try:
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
        rows: list[list[Any]] = []
        decisions: list[dict[str, Any]] = []
        cumulative = 0.0
        timestamp = datetime.now().isoformat(timespec="seconds")

        for source_row, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(value is None for value in row):
                continue
            date_value = row[0] if len(row) > 0 else ""
            old_category = row[1] if len(row) > 1 else ""
            old_amount = _amount(row[2] if len(row) > 2 else 0)
            description = str(row[4] or "") if len(row) > 4 else ""
            # v1.1.0 had a single Timestamp column at index 5. New files have
            # Created Timestamp and Last Updated Timestamp at indices 5 and 6.
            created_at = str(row[5] or "") if len(row) > 5 else ""
            updated_at = str(row[6] or "") if len(row) > 6 else created_at
            updated_device = str(row[7] or "legacy") if len(row) > 7 else "legacy"
            new_category, sign, reason = _map_legacy_category(old_category, description)
            new_amount = abs(old_amount) * sign if sign else old_amount
            cumulative += new_amount
            rows.append([date_value, new_category, new_amount, cumulative, description, created_at or timestamp, updated_at or timestamp, updated_device])
            decisions.append({
                "source_row": source_row,
                "date": str(date_value or ""),
                "original_category": str(old_category or ""),
                "original_amount": old_amount,
                "description": description,
                "migrated_category": new_category,
                "migrated_amount": new_amount,
                "decision": reason,
            })
        return rows, decisions
    finally:
        workbook.close()


def _write_and_validate(path: Path, rows: list[list[Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()

    validated_workbook = load_workbook(path, data_only=True)
    try:
        validated = validated_workbook[SHEET_NAME]
        actual_headers = [validated.cell(1, index).value for index in range(1, len(HEADERS) + 1)]
        actual_rows = sum(1 for row in validated.iter_rows(min_row=2, values_only=True) if any(value is not None for value in row))
        if actual_headers != HEADERS or actual_rows != len(rows):
            raise ValueError("Staged migration workbook validation failed.")
    finally:
        validated_workbook.close()


def migrate_bank_workbook(data_dir: Path, *, apply: bool) -> dict[str, Any]:
    source = data_dir / "bank_transactions.xlsx"
    if not source.exists():
        return {"status": "not-found", "source_file": str(source), "migration": MIGRATION_VERSION}

    rows, decisions = _read_and_map(source)
    mapped = [item for item in decisions if item["decision"] != "already compatible"]
    report: dict[str, Any] = {
        "migration": MIGRATION_VERSION,
        "mode": "apply" if apply else "preview",
        "source_file": str(source),
        "row_count": len(decisions),
        "mapped_row_count": len(mapped),
        "unchanged_row_count": len(decisions) - len(mapped),
        "decisions": decisions,
    }
    if not apply:
        return report

    tag = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_dir = data_dir / "backups" / f"v1.1.0-to-{MIGRATION_VERSION}" / tag
    backup_dir.mkdir(parents=True, exist_ok=False)
    backup_file = backup_dir / source.name
    staged_workbook = data_dir / f".bank_transactions.{MIGRATION_VERSION}-{tag}.xlsx"
    staged_report = data_dir / f".bank_migration_report_{MIGRATION_VERSION}-{tag}.json"
    report_file = backup_dir / "migration-report.json"

    try:
        # Preserve raw data before any migration file is created.
        shutil.copy2(source, backup_file)
        _write_and_validate(staged_workbook, rows)
        report.update({"status": "migrated", "backup_folder": str(backup_dir), "backup_file": str(backup_file), "report_file": str(report_file)})
        staged_report.write_text(json.dumps(report, indent=2), encoding="utf-8")
        os.replace(staged_workbook, source)
        os.replace(staged_report, report_file)
        return report
    finally:
        for path in (staged_workbook, staged_report):
            if path.exists():
                path.unlink()


def _migrate_share_rows(source: Path) -> list[list[Any]]:
    workbook = load_workbook(source, data_only=True)
    try:
        sheet = workbook["Share"] if "Share" in workbook.sheetnames else workbook.active
        source_headers = [str(cell.value or "").strip() for cell in sheet[1]]
        rows: list[list[Any]] = []
        timestamp = datetime.now().isoformat(timespec="seconds")
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(value is None for value in row):
                continue
            created_at = str(_value_for_header(row, source_headers, "Created Timestamp") or _value_for_header(row, source_headers, "Timestamp") or "")
            last_updated_at = str(_value_for_header(row, source_headers, "Last Updated Timestamp") or created_at or "")
            updated_device = str(_value_for_header(row, source_headers, "Updated Device") or "legacy")
            rows.append([
                row[0] if len(row) > 0 else "",
                row[1] if len(row) > 1 else "",
                row[2] if len(row) > 2 else "",
                row[3] if len(row) > 3 else "",
                row[4] if len(row) > 4 else 0,
                row[5] if len(row) > 5 else 0,
                row[6] if len(row) > 6 else "",
                row[7] if len(row) > 7 else "",
                row[8] if len(row) > 8 else "",
                row[9] if len(row) > 9 else 0,
                created_at or timestamp,
                last_updated_at or created_at or timestamp,
                updated_device,
            ])
        return rows
    finally:
        workbook.close()


def _migrate_personal_finance_rows(source: Path, sheet_name: str) -> list[list[Any]]:
    workbook = load_workbook(source, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        source_headers = [str(cell.value or "").strip() for cell in sheet[1]]
        rows: list[list[Any]] = []
        timestamp = datetime.now().isoformat(timespec="seconds")
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(value is None for value in row):
                continue
            created_at = str(_value_for_header(row, source_headers, "Created Timestamp") or _value_for_header(row, source_headers, "Timestamp") or "")
            last_updated_at = str(_value_for_header(row, source_headers, "Last Updated Timestamp") or created_at or "")
            source_ref = _value_for_header(row, source_headers, "Source Ref") or ""
            updated_device = str(_value_for_header(row, source_headers, "Updated Device") or "legacy")
            rows.append([
                row[0] if len(row) > 0 else "",
                row[1] if len(row) > 1 else "",
                row[2] if len(row) > 2 else "",
                row[3] if len(row) > 3 else "",
                row[4] if len(row) > 4 else 0,
                row[5] if len(row) > 5 else 0,
                row[6] if len(row) > 6 else "",
                row[7] if len(row) > 7 else "manual",
                created_at or timestamp,
                last_updated_at or created_at or timestamp,
                source_ref or "",
                updated_device,
            ])
        return rows
    finally:
        workbook.close()


def _value_for_header(row: tuple[Any, ...], headers: list[str], header: str) -> Any:
    try:
        index = headers.index(header)
    except ValueError:
        return None
    return row[index] if index < len(row) else None


def _write_rows(path: Path, sheet_name: str, headers: list[str], rows: list[list[Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def _migrate_header_workbook(data_dir: Path, source: Path, sheet_name: str, headers: list[str], rows: list[list[Any]], label: str) -> dict[str, Any]:
    tag = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_dir = data_dir / "backups" / label / tag
    backup_dir.mkdir(parents=True, exist_ok=False)
    backup_file = backup_dir / source.name
    staged_workbook = data_dir / f".{source.stem}.{label}-{tag}.xlsx"
    try:
        shutil.copy2(source, backup_file)
        _write_rows(staged_workbook, sheet_name, headers, rows)
        os.replace(staged_workbook, source)
        return {"status": "migrated", "source_file": str(source), "backup_file": str(backup_file), "row_count": len(rows)}
    finally:
        if staged_workbook.exists():
            staged_workbook.unlink()


def migrate_share_workbook(data_dir: Path) -> dict[str, Any]:
    source = data_dir / "share_transactions.xlsx"
    if not source.exists():
        return {"status": "not-found", "source_file": str(source), "migration": MIGRATION_VERSION}
    if not _workbook_needs_header_migration(source, "Share", SHARE_HEADERS):
        return {"status": "not-needed", "source_file": str(source), "migration": MIGRATION_VERSION}
    return _migrate_header_workbook(data_dir, source, "Share", SHARE_HEADERS, _migrate_share_rows(source), "share-updated-device")


def migrate_personal_finance_workbook(data_dir: Path, file_name: str, sheet_name: str) -> dict[str, Any]:
    source = data_dir / file_name
    if not source.exists():
        return {"status": "not-found", "source_file": str(source), "migration": MIGRATION_VERSION}
    if not _workbook_needs_header_migration(source, sheet_name, PERSONAL_FINANCE_HEADERS):
        return {"status": "not-needed", "source_file": str(source), "migration": MIGRATION_VERSION}
    return _migrate_header_workbook(data_dir, source, sheet_name, PERSONAL_FINANCE_HEADERS, _migrate_personal_finance_rows(source, sheet_name), f"{source.stem}-updated-device")


def run_pending_data_migrations(data_dir: Path | None = None) -> dict[str, Any]:
    target_dir = data_dir or get_data_dir()
    source = target_dir / "bank_transactions.xlsx"
    results: list[dict[str, Any]] = []
    if source.exists() and _bank_needs_migration(source):
        results.append(migrate_bank_workbook(target_dir, apply=True))
    else:
        results.append({"status": "not-found" if not source.exists() else "not-needed", "migration": MIGRATION_VERSION, "source_file": str(source)})
    results.append(migrate_share_workbook(target_dir))
    results.append(migrate_personal_finance_workbook(target_dir, "personal_finance_bank_flow.xlsx", "Bank Flow"))
    results.append(migrate_personal_finance_workbook(target_dir, "personal_finance_cash_flow.xlsx", "Cash Flow"))
    migrated = [result for result in results if result.get("status") == "migrated"]
    return {
        "status": "migrated" if migrated else "not-needed",
        "migration": MIGRATION_VERSION,
        "source_file": str(target_dir),
        "results": results,
    }
