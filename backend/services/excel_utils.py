import logging
import shutil
import threading
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)
_excel_lock = threading.Lock()
LAST_RECOVERY_WARNINGS: dict[str, str] = {}


# ---------------------------------------------------------------------------
# Numeric helpers
# ---------------------------------------------------------------------------

def to_float(value: object) -> float:
    """Coerce any value to float, returning 0.0 on failure."""
    try:
        return float(value or 0.0)
    except (TypeError, ValueError):
        return 0.0


def to_int(value: object) -> int:
    """Coerce any value to int, returning 0 on failure."""
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


# ---------------------------------------------------------------------------
# Timestamp helpers
# ---------------------------------------------------------------------------

def current_timestamp() -> str:
    """ISO-8601 timestamp at seconds precision — used by Share and PF services."""
    return datetime.now().isoformat(timespec="seconds")


def current_timestamp_us() -> str:
    """ISO-8601 timestamp at microseconds precision — used by Bank service
    so that two rapid edits on the same record are always distinguishable."""
    return datetime.now().isoformat(timespec="microseconds")


# ---------------------------------------------------------------------------
# Record-ID / Excel-row validation
# ---------------------------------------------------------------------------

def validate_record_id(sheet, record_id: int) -> int:
    """
    Convert a public record_id (0-based, header excluded) to a 1-based Excel row.
    Raises ValueError if out of range so callers get HTTP 400.
    """
    excel_row = record_id + 1          # +1 because row 1 is the header
    if excel_row < 2 or excel_row > sheet.max_row:
        raise ValueError(
            f"record_id {record_id} is out of range "
            f"(valid: 0–{sheet.max_row - 2})."
        )
    return excel_row


# ---------------------------------------------------------------------------
# Standard API response shape
# ---------------------------------------------------------------------------

def api_response(message: str, data: Any = None) -> dict:
    """Return the standard {message, data} envelope used by all route handlers."""
    return {"message": message, "data": data}


# ---------------------------------------------------------------------------
# Corrupted file recovery & safe workbook loader
# ---------------------------------------------------------------------------

def find_most_recent_valid_backup(file_path: Path) -> Path | None:
    """
    Search file_path.parent/backups/ (excluding backups/corrupted/) for the
    newest uncorrupted backup file matching file_path.name.
    """
    backups_dir = file_path.parent / "backups"
    if not backups_dir.exists():
        return None

    candidates: list[Path] = []
    for p in backups_dir.glob(f"**/{file_path.name}"):
        # Exclude already-broken copies in backups/corrupted/
        if "corrupted" in p.parts:
            continue
        if p.resolve() == file_path.resolve():
            continue
        candidates.append(p)

    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)

    for candidate in candidates:
        try:
            wb = load_workbook(candidate, read_only=True, data_only=True)
            _ = wb.active.max_row
            wb.close()
            return candidate
        except Exception:
            continue

    return None


def repair_or_recover_workbook(
    file_path: Path,
    sheet_name: str,
    headers: list[str],
) -> Workbook:
    """
    When a workbook is missing or corrupted:
    1. Quarantine corrupted file to backups/corrupted/<timestamp>/ if present.
    2. Attempt automatic restoration from the newest valid backup.
    3. If no valid backup exists, create a fresh empty workbook and set a warning.
    """
    if file_path.exists() and file_path.stat().st_size > 0:
        try:
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
            quarantine_dir = file_path.parent / "backups" / "corrupted" / timestamp
            quarantine_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(file_path, quarantine_dir / file_path.name)
            logger.warning(f"Corrupted file {file_path} quarantined to {quarantine_dir}")
        except Exception as exc:
            logger.error(f"Could not quarantine corrupted file {file_path}: {exc}")

    # Search for an uncorrupted backup
    valid_backup = find_most_recent_valid_backup(file_path)
    if valid_backup is not None:
        try:
            file_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(valid_backup, file_path)
            logger.warning(
                f"File {file_path.name} was corrupted or missing. "
                f"Automatically restored from valid backup: {valid_backup}"
            )
            return load_workbook(file_path)
        except Exception as exc:
            logger.error(f"Failed to restore from valid backup {valid_backup}: {exc}")

    # No valid backup exists -> create new empty file and log warning
    warning_msg = (
        f"{file_path.name} was corrupted and no valid backup was found; "
        f"a new empty file was created."
    )
    logger.error(warning_msg)
    LAST_RECOVERY_WARNINGS[file_path.name] = warning_msg

    try:
        file_path.unlink(missing_ok=True)
    except Exception:
        pass

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    wb.save(file_path)
    return wb


def safe_load_workbook(
    file_path: Path,
    sheet_name: str,
    headers: list[str],
    data_only: bool = False,
    read_only: bool = False,
) -> Workbook:
    """
    Safely load a workbook. If missing, 0-byte, or corrupted, automatically
    recovers from the newest valid backup before attempting to load.
    """
    if not file_path.exists() or file_path.stat().st_size == 0:
        return repair_or_recover_workbook(file_path, sheet_name, headers)

    try:
        return load_workbook(file_path, data_only=data_only, read_only=read_only)
    except (zipfile.BadZipFile, InvalidFileException, OSError, KeyError) as err:
        logger.warning(f"File {file_path} is corrupted ({err}). Attempting backup recovery.")
        return repair_or_recover_workbook(file_path, sheet_name, headers)


# ---------------------------------------------------------------------------
# Generic workbook initialiser
# ---------------------------------------------------------------------------

def ensure_workbook_exists(
    file_path: Path,
    sheet_name: str,
    headers: list[str],
    data_dir: Path | None = None,
) -> None:
    """
    Ensure *file_path* exists and its first sheet has the expected *headers*.
    Only saves if headers or sheet structure were modified.
    """
    with _excel_lock:
        if data_dir is not None:
            data_dir.mkdir(parents=True, exist_ok=True)

        wb = safe_load_workbook(file_path, sheet_name, headers)
        try:
            modified = False
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(headers)
                modified = True
            else:
                ws = wb[sheet_name]
                existing_headers = [ws.cell(1, col).value for col in range(1, len(headers) + 1)]
                if existing_headers != headers:
                    for col, h in enumerate(headers, start=1):
                        ws.cell(row=1, column=col, value=h)
                    modified = True

            if modified:
                wb.save(file_path)
        finally:
            wb.close()

