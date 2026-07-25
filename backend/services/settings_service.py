"""Settings import/export helpers for FinLedge data workbooks."""

from __future__ import annotations

import os
import shutil
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from ..models import (
    BankCategory,
    ShareCategory,
    PF_EXPENSE_CATEGORIES,
    PF_INCOME_CATEGORIES,
)
from . import bank_service, personal_finance_service, share_service
from .path_utils import get_data_dir

DATA_DIR = get_data_dir()

# ---------------------------------------------------------------------------
# Column aliases: map old / alternate names → canonical current header name.
# Keys are lowercase; values must match the canonical header exactly.
# ---------------------------------------------------------------------------
_COLUMN_ALIASES: dict[str, str] = {
    # Bank Services — old versions called it "Cumulative Total"
    "cumulative total": "Cumulative Amount",
    "cumulative balance": "Cumulative Amount",
    "running total": "Cumulative Amount",
    # Share Portfolio — some exports used "Timestamp" for the row timestamp
    "created at": "Timestamp",
    "created timestamp": "Timestamp",
    # Personal Finance — alternate Direction names
    "type": "Direction",
    "transaction type": "Direction",
    # Generic
    "note": "Description",
    "notes": "Description",
    "remarks": "Description",
}

# ---------------------------------------------------------------------------
# Truly required columns per data type (minimum a user export must have).
# Derived / calculated / internal columns are all optional and auto-filled.
# ---------------------------------------------------------------------------
_BANK_REQUIRED = ["Date", "Category", "Amount"]
_SHARE_REQUIRED = ["Date", "Share Name", "Category", "Per Unit Price", "Allotted", "Buy/Sell"]
_PF_REQUIRED = ["Date", "Direction", "Category", "Amount"]

# ---------------------------------------------------------------------------
# Category migration / normalisation tables
# ---------------------------------------------------------------------------

# Bank: legacy categories that need re-mapping (same rules as data_migration_service.py)
_BANK_LEGACY_CATEGORIES = {"income", "service cost", "investment cost", "operation cost"}
_BANK_RETIRED_CATEGORIES = {"atm charge", "sms charge"}
_BANK_VALID_CATEGORIES = {c.value for c in BankCategory}
_BANK_INCOME_CATEGORIES = {"Interest Earned"}

# Share: any casing / verbose label → canonical lowercase key stored in the workbook
_SHARE_CATEGORY_ALIASES: dict[str, str] = {
    "ipo": "ipo",
    "ipo entry": "ipo",
    "initial public offering": "ipo",
    "sip": "sip",
    "sip investment": "sip",
    "systematic investment plan": "sip",
    "buy": "buy",
    "secondary buy": "buy",
    "purchase": "buy",
    "sell": "sell",
    "sell shares": "sell",
    "sale": "sell",
    "dividend": "dividend",
    "dividend (cash)": "dividend",
    "cash dividend": "dividend",
    "dividend (bonus)": "dividend",
    "bonus dividend": "dividend",
    "bonus": "dividend",
}
_SHARE_VALID_CATEGORIES = {c.value for c in ShareCategory}

# PF direction aliases: any variant → "income" or "expense"
_PF_DIRECTION_ALIASES: dict[str, str] = {
    "income": "income",
    "credit": "income",
    "in": "income",
    "inflow": "income",
    "expense": "expense",
    "expenses": "expense",
    "debit": "expense",
    "out": "expense",
    "outflow": "expense",
    "cost": "expense",
    "charge": "expense",
    "spend": "expense",
    "spending": "expense",
}

# PF expense category aliases: old/variant → current valid label
_PF_EXPENSE_ALIASES: dict[str, str] = {
    "food & dining": "Food",
    "food and dining": "Food",
    "dining": "Food",
    "restaurant": "Food",
    "grocery": "Food",
    "groceries": "Food",
    "transport": "Transportation",
    "commute": "Transportation",
    "fuel": "Transportation",
    "petrol": "Transportation",
    "movie": "Entertainment",
    "movies": "Entertainment",
    "recreation": "Entertainment",
    "clothing": "Shopping",
    "clothes": "Shopping",
    "medical": "Health",
    "hospital": "Health",
    "medicine": "Health",
    "school": "Education",
    "university": "Education",
    "tuition": "Education",
    "utility": "Bills",
    "utilities": "Bills",
    "electricity": "Bills",
    "internet": "Bills",
    "water": "Bills",
    "phone": "Bills",
    "mobile": "Bills",
    "house rent": "Rent",
    "apartment rent": "Rent",
    "vacation": "Travel",
    "trip": "Travel",
    "invest": "Investment",
    "investment cost": "Investment",
    "share market": "Share Market",
    "stock": "Share Market",
    "sip": "SIP",
}
_PF_EXPENSE_VALID = {c.value for c in PF_EXPENSE_CATEGORIES}

# PF income category aliases
_PF_INCOME_ALIASES: dict[str, str] = {
    "wages": "Salary",
    "monthly salary": "Salary",
    "bonus salary": "Salary",
    "part time": "Freelance",
    "part-time": "Freelance",
    "contract work": "Freelance",
    "business income": "Business",
    "profit": "Business",
    "lottery": "Prize/Lottery",
    "prize": "Prize/Lottery",
    "gift money": "Gift",
    "present": "Gift",
    "cashback": "Refund",
    "cash back": "Refund",
    "return": "Refund",
    "investment return": "Investment Return",
    "investment income": "Investment Income",
    "capital gain": "Investment Return",
    "interest income": "Investment Income",
    "dividend income": "Dividend",
    "dividend": "Dividend",
    "share proceeds": "Share Sell Proceeds",
    "sell proceeds": "Share Sell Proceeds",
    "share sell": "Share Sell Proceeds",
}
_PF_INCOME_VALID = {c.value for c in PF_INCOME_CATEGORIES}


def _migrate_bank_category(category: str, description: str) -> tuple[str, int]:
    """
    Map a bank category to the current v1.2.0 category set.
    Returns (new_category, sign) where sign is +1 for income, -1 for charge, 0 = keep amount as-is.
    Mirrors the logic in data_migration_service._map_legacy_category exactly.
    """
    orig = category.strip().lower()
    text = description.strip().lower()
    compact = text.replace(" ", "")

    # Already a valid current category — keep it
    if category.strip() in _BANK_VALID_CATEGORIES:
        sign = 1 if category.strip() in _BANK_INCOME_CATEGORIES else -1
        return category.strip(), 0  # amount already signed correctly

    if orig == "atm charge":
        return "Debit Card Charge", -1
    if orig == "sms charge":
        return "Mobile Banking Charge", -1
    if orig not in _BANK_LEGACY_CATEGORIES:
        # Unknown category — put under Other Charges
        return "Other Charges", -1
    if orig == "income" and ("interest" in text or "int" in text):
        return "Interest Earned", 1
    if orig == "service cost" and "tax" in text:
        return "Interest Tax", -1
    if orig == "service cost" and "mobile banking" in text and "renew" in text:
        return "Mobile Banking Charge", -1
    if orig == "service cost" and "bank" in text and "renew" in text:
        return "Mobile Banking Charge", -1
    if orig == "service cost" and "mobile" in text:
        return "Mobile Banking Charge", -1
    if orig == "service cost" and "card" in text and "install" in text:
        return "Debit Card Charge", -1
    if orig == "investment cost" and "demat" in text and "meroshare" in compact and "renew" in text:
        return "Demat & MeroShare Renewal", -1
    if orig == "investment cost" and "meroshare" in compact and "renew" in text:
        return "MeroShare Renewal", -1
    if orig == "investment cost" and "demat" in text and "renew" in text:
        return "Demat Renewal", -1
    if orig == "investment cost" and "broker" in text and "renew" in text:
        return "Broker Renewal", -1
    # income that doesn't match interest → keep as income category
    if orig == "income":
        return "Other Charges", 1  # positive amount, unknown income
    return "Other Charges", -1


def _migrate_share_category(category: str) -> str:
    """Normalise a Share Portfolio category to the current lowercase key."""
    key = category.strip().lower()
    return _SHARE_CATEGORY_ALIASES.get(key, key if key in _SHARE_VALID_CATEGORIES else "buy")


def _migrate_pf_direction(direction: str) -> str:
    """Normalise a PF direction value to 'income' or 'expense'."""
    key = direction.strip().lower()
    return _PF_DIRECTION_ALIASES.get(key, "expense")


def _migrate_pf_category(category: str, direction: str) -> str:
    """Map a PF category to the closest current valid category for the given direction."""
    raw = category.strip()
    key = raw.lower()
    if direction == "income":
        if raw in _PF_INCOME_VALID:
            return raw
        return _PF_INCOME_ALIASES.get(key, "Other Income")
    else:
        if raw in _PF_EXPENSE_VALID:
            return raw
        return _PF_EXPENSE_ALIASES.get(key, "Other")


DATA_TYPE_CONFIG: dict[str, dict[str, Any]] = {
    "bank": {
        "label": "Bank Services",
        "filename": "bank_transactions.xlsx",
        "path": bank_service.FILE_PATH,
        "headers": bank_service.HEADERS,
        "required_headers": _BANK_REQUIRED,
        "sheet_name": bank_service.SHEET_NAME,
    },
    "share": {
        "label": "Share Portfolio",
        "filename": "share_transactions.xlsx",
        "path": share_service.FILE_PATH,
        "headers": share_service.HEADERS,
        "required_headers": _SHARE_REQUIRED,
        "sheet_name": share_service.SHEET_NAME,
    },
    "pf-bank": {
        "label": "Personal Expenses — Bank Flow",
        "filename": "personal_finance_bank_flow.xlsx",
        "path": personal_finance_service.BANK_FLOW_FILE_PATH,
        "headers": personal_finance_service.HEADERS,
        "required_headers": _PF_REQUIRED,
        "sheet_name": personal_finance_service.SHEET_NAME,
    },
    "pf-cash": {
        "label": "Personal Expenses — Cash Flow",
        "filename": "personal_finance_cash_flow.xlsx",
        "path": personal_finance_service.CASH_FLOW_FILE_PATH,
        "headers": personal_finance_service.HEADERS,
        "required_headers": _PF_REQUIRED,
        "sheet_name": personal_finance_service.SHEET_NAME,
    },
}


def list_data_types() -> list[dict[str, Any]]:
    return [
        {
            "id": dt,
            "label": cfg["label"],
            "filename": cfg["filename"],
            "headers": list(cfg["headers"]),
            "required_headers": list(cfg["required_headers"]),
        }
        for dt, cfg in DATA_TYPE_CONFIG.items()
    ]


def get_data_type_config(data_type: str) -> dict[str, Any]:
    normalized = str(data_type or "").strip().lower()
    config = DATA_TYPE_CONFIG.get(normalized)
    if not config:
        supported = ", ".join(DATA_TYPE_CONFIG)
        raise ValueError(f"Unknown data type '{data_type}'. Supported types: {supported}.")
    return config


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _to_float(value: object) -> float:
    try:
        return float(str(value).replace(",", "").strip()) if value not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def _noon_timestamp(date_value: Any, fallback_ts: str) -> str:
    """Return a noon-time timestamp string from a cell date value, or fallback_ts."""
    if date_value is None:
        return fallback_ts
    # date_value may be a datetime, date object, or an ISO string
    try:
        if isinstance(date_value, datetime):
            return date_value.replace(hour=12, minute=0, second=0, microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
        if hasattr(date_value, "year"):  # date object
            return f"{date_value.isoformat()} 12:00:00"
        # Try parsing string
        s = str(date_value).strip()[:10]  # take YYYY-MM-DD portion
        return f"{s} 12:00:00"
    except Exception:
        return fallback_ts


def _apply_aliases(headers: list[str]) -> list[str]:
    """Normalise column headers: apply aliases so old names map to current canonical names."""
    result = []
    for h in headers:
        lower = h.strip().lower()
        result.append(_COLUMN_ALIASES.get(lower, h.strip()))
    return result


def _read_workbook_rows(
    source_path: Path,
    sheet_name: str,
) -> tuple[list[str], list[list[Any]]]:
    """Return (header_row, data_rows) from the workbook. Active sheet used if sheet_name not found."""
    wb = load_workbook(source_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return [], []

    raw_headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    headers = _apply_aliases(raw_headers)
    data = [list(r) for r in rows[1:] if any(v is not None for v in r)]
    return headers, data


def _validate_required_headers(actual_headers: list[str], required_headers: list[str]) -> list[str]:
    """Return list of missing required headers (empty = OK). Case-insensitive."""
    actual_lower = {h.lower() for h in actual_headers}
    return [h for h in required_headers if h.lower() not in actual_lower]


# ---------------------------------------------------------------------------
# Per-data-type row remappers with derived-column calculation
# ---------------------------------------------------------------------------

def _remap_bank_rows(
    rows: list[list[Any]],
    actual_headers: list[str],
    full_headers: list[str],
    import_ts: str,
    initial_cumulative: float = 0.0,
) -> list[list[Any]]:
    """
    Map bank import rows into the canonical Bank HEADERS order.
    - Category is migrated from old v1.1.0 names to the current v1.2.0 category set.
    - Amount sign is corrected if the category migration requires it (income = +, charge = -).
    - Cumulative Amount is recalculated as a fresh running sum.
    - Description defaults to "".
    - Timestamps default to noon on the row's own date.
    """
    idx = {h.lower(): i for i, h in enumerate(actual_headers)}
    running_cumulative = initial_cumulative
    result = []

    for row in rows:
        def get(col: str, default: Any = None) -> Any:
            i = idx.get(col.lower())
            return row[i] if i is not None else default

        date_val = get("Date", "")
        raw_category = str(get("Category", "") or "").strip()
        amount = _to_float(get("Amount", 0))
        description = str(get("Description", "") or "").strip()

        # Migrate category and fix amount sign if needed
        category, sign = _migrate_bank_category(raw_category, description)
        if sign == 1:
            amount = abs(amount)   # income must be positive
        elif sign == -1:
            amount = -abs(amount)  # charge must be negative
        # sign == 0 means already a valid category — keep amount as-is

        # Recalculate running cumulative from the (now-signed) amount
        running_cumulative = round(running_cumulative + amount, 10)

        ts = _noon_timestamp(date_val, import_ts)
        created_ts = str(get("Created Timestamp", ts) or ts)
        updated_ts = str(get("Last Updated Timestamp", ts) or ts)

        remapped = []
        for col in full_headers:
            col_l = col.lower()
            if col_l == "date":
                remapped.append(str(date_val).strip()[:10] if date_val else "")
            elif col_l == "category":
                remapped.append(category)
            elif col_l == "amount":
                remapped.append(amount)
            elif col_l == "cumulative amount":
                remapped.append(round(running_cumulative, 2))
            elif col_l == "description":
                remapped.append(description)
            elif col_l == "created timestamp":
                remapped.append(created_ts)
            elif col_l == "last updated timestamp":
                remapped.append(updated_ts)
            else:
                remapped.append(None)
        result.append(remapped)

    return result


def _remap_share_rows(
    rows: list[list[Any]],
    actual_headers: list[str],
    full_headers: list[str],
    import_ts: str,
) -> list[list[Any]]:
    """
    Map share import rows into canonical Share HEADERS order.
    - Category is normalised to current lowercase key (ipo/sip/buy/sell/dividend).
    - Profit/Loss and Cumulative Profit are preserved from the file if present,
      otherwise set to None.
    - ASBA Charge defaults to 0.
    - Total Amount: calculated as Per Unit Price × Allotted if not supplied.
    - Timestamp defaults to noon on the row's own date.
    - Sync Ref defaults to None.
    """
    idx = {h.lower(): i for i, h in enumerate(actual_headers)}
    result = []

    for row in rows:
        def get(col: str, default: Any = None) -> Any:
            i = idx.get(col.lower())
            return row[i] if i is not None else default

        date_val = get("Date", "")
        share_name = str(get("Share Name", "") or "").strip()
        raw_category = str(get("Category", "") or "").strip()
        # Migrate category to current canonical value
        category = _migrate_share_category(raw_category)
        per_unit = _to_float(get("Per Unit Price", 0))
        asba = _to_float(get("ASBA Charge", 0))
        allotted = _to_float(get("Allotted", 0))
        buy_sell = str(get("Buy/Sell", "") or "").strip()

        # Total Amount: use existing value, calculate from price×qty if absent
        raw_total = get("Total Amount")
        total_amount = _to_float(raw_total) if raw_total is not None else round(per_unit * allotted, 2)

        profit_loss = get("Profit/Loss")
        cumulative_profit = get("Cumulative Profit")

        ts = _noon_timestamp(date_val, import_ts)
        row_ts = str(get("Timestamp", ts) or ts)
        sync_ref = get("Sync Ref", None)

        remapped = []
        for col in full_headers:
            col_l = col.lower()
            if col_l == "date":
                remapped.append(str(date_val).strip()[:10] if date_val else "")
            elif col_l == "share name":
                remapped.append(share_name)
            elif col_l == "category":
                remapped.append(category)
            elif col_l == "per unit price":
                remapped.append(per_unit)
            elif col_l == "asba charge":
                remapped.append(asba)
            elif col_l == "allotted":
                remapped.append(int(allotted) if allotted == int(allotted) else allotted)
            elif col_l == "buy/sell":
                remapped.append(buy_sell)
            elif col_l == "total amount":
                remapped.append(total_amount)
            elif col_l == "profit/loss":
                remapped.append(_to_float(profit_loss) if profit_loss is not None else None)
            elif col_l == "cumulative profit":
                remapped.append(_to_float(cumulative_profit) if cumulative_profit is not None else None)
            elif col_l == "timestamp":
                remapped.append(row_ts)
            elif col_l == "sync ref":
                remapped.append(sync_ref)
            else:
                remapped.append(None)
        result.append(remapped)

    return result


def _remap_pf_rows(
    rows: list[list[Any]],
    actual_headers: list[str],
    full_headers: list[str],
    import_ts: str,
    flow_type_override: str = "",
) -> list[list[Any]]:
    """
    Map personal-finance import rows into canonical Personal Finance HEADERS order.
    - Direction is normalised to 'income' or 'expense' (handles credit/debit/in/out etc.).
    - Category is migrated to the closest current valid category for the direction.
    - Signed Amount is recalculated from the normalised direction.
    - Flow Type defaults to flow_type_override or 'bank'.
    - Source defaults to 'manual'.
    - Timestamp defaults to noon on the row's own date.
    """
    idx = {h.lower(): i for i, h in enumerate(actual_headers)}
    result = []

    for row in rows:
        def get(col: str, default: Any = None) -> Any:
            i = idx.get(col.lower())
            return row[i] if i is not None else default

        date_val = get("Date", "")
        flow_type = str(get("Flow Type", flow_type_override or "bank") or flow_type_override or "bank").strip().lower()
        raw_direction = str(get("Direction", "expense") or "expense").strip()
        # Normalise direction to 'income' or 'expense'
        direction = _migrate_pf_direction(raw_direction)
        raw_category = str(get("Category", "") or "").strip()
        # Migrate category to the current valid set for this direction
        category = _migrate_pf_category(raw_category, direction)
        amount = abs(_to_float(get("Amount", 0)))
        description = str(get("Description", "") or "").strip()
        source = str(get("Source", "manual") or "manual").strip()
        source_ref = get("Source Ref", None)

        # Always recalculate signed amount from the normalised direction
        signed_amount = amount if direction == "income" else -amount

        ts = _noon_timestamp(date_val, import_ts)
        row_ts = str(get("Timestamp", ts) or ts)

        remapped = []
        for col in full_headers:
            col_l = col.lower()
            if col_l == "date":
                remapped.append(str(date_val).strip()[:10] if date_val else "")
            elif col_l == "flow type":
                remapped.append(flow_type)
            elif col_l == "direction":
                remapped.append(direction)
            elif col_l == "category":
                remapped.append(category)
            elif col_l == "amount":
                remapped.append(amount)
            elif col_l == "signed amount":
                remapped.append(signed_amount)
            elif col_l == "description":
                remapped.append(description)
            elif col_l == "source":
                remapped.append(source)
            elif col_l == "timestamp":
                remapped.append(row_ts)
            elif col_l == "source ref":
                remapped.append(source_ref)
            else:
                remapped.append(None)
        result.append(remapped)

    return result


def _remap_rows(
    data_type: str,
    rows: list[list[Any]],
    actual_headers: list[str],
    full_headers: list[str],
    import_ts: str,
    initial_cumulative: float = 0.0,
    flow_type_override: str = "",
) -> list[list[Any]]:
    if data_type == "bank":
        return _remap_bank_rows(rows, actual_headers, full_headers, import_ts, initial_cumulative)
    elif data_type == "share":
        return _remap_share_rows(rows, actual_headers, full_headers, import_ts)
    else:
        return _remap_pf_rows(rows, actual_headers, full_headers, import_ts, flow_type_override)


# ---------------------------------------------------------------------------
# Backup / live-row helpers
# ---------------------------------------------------------------------------

def _backup_live_file(live_path: Path) -> Path | None:
    if not live_path.exists():
        return None
    tag = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    backup_dir = DATA_DIR / "backups" / "import" / tag
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_file = backup_dir / live_path.name
    shutil.copy2(live_path, backup_file)
    return backup_file


def _count_live_rows(live_path: Path, sheet_name: str) -> int:
    if not live_path.exists():
        return 0
    try:
        wb = load_workbook(live_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            count = sum(
                1 for row in ws.iter_rows(min_row=2, values_only=True)
                if any(v is not None for v in row)
            )
        finally:
            wb.close()
        return count
    except Exception:
        return 0


def _read_live_rows(live_path: Path, sheet_name: str, data_type: str = "") -> tuple[list[list[Any]], float]:
    """Read existing live data rows (in canonical column order).
    Also returns the last cumulative amount (bank only, column index 3)."""
    if not live_path.exists():
        return [], 0.0
    _, rows = _read_workbook_rows(live_path, sheet_name)
    last_cumulative = 0.0
    if rows and data_type == "bank":
        try:
            last_cumulative = float(rows[-1][3] or 0)
        except (IndexError, TypeError, ValueError):
            last_cumulative = 0.0
    return rows, last_cumulative


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def import_data_file(
    data_type: str,
    uploaded_path: Path,
    mode: str = "replace",
) -> dict[str, Any]:
    """
    Import an Excel file for the given data_type.

    mode:
      "replace" — discard current live data and replace with the import file.
      "merge"   — keep existing live rows, append imported rows after them.

    Old column names are automatically aliased to their current equivalents.
    Derived columns (Cumulative Amount, Signed Amount, Total Amount) are
    calculated from the available data if not present in the import file.
    Internal timestamp columns default to noon on each row's own date.
    """
    config = get_data_type_config(data_type)
    live_path = Path(config["path"])
    full_headers: list[str] = list(config["headers"])
    required_headers: list[str] = list(config["required_headers"])
    sheet_name: str = str(config["sheet_name"])

    if not uploaded_path.exists():
        raise ValueError("Import file was not found on the server.")

    suffix = uploaded_path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Import file must be an Excel workbook (.xlsx).")

    try:
        actual_headers, import_rows = _read_workbook_rows(uploaded_path, sheet_name)
    except Exception as exc:
        raise ValueError(f"Could not read the import workbook: {exc}") from exc

    if not actual_headers:
        raise ValueError("The import file appears to be empty — no header row found.")

    # Validate that all truly required columns are present (aliases already applied)
    missing = _validate_required_headers(actual_headers, required_headers)
    if missing:
        raise ValueError(
            f"The import file is missing required columns: {', '.join(missing)}. "
            f"Required: {', '.join(required_headers)}"
        )

    import_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # For merge mode: load existing live rows first
    live_rows: list[list[Any]] = []
    initial_cumulative = 0.0
    flow_type_override = "bank" if data_type == "pf-bank" else ("cash" if data_type == "pf-cash" else "")

    if mode == "merge":
        live_rows, initial_cumulative = _read_live_rows(live_path, sheet_name, data_type)
    else:
        initial_cumulative = 0.0

    # For bank in replace mode: cumulative starts from 0.
    # For bank in merge mode: cumulative continues from last live row.
    remapped = _remap_rows(
        data_type,
        import_rows,
        actual_headers,
        full_headers,
        import_ts,
        initial_cumulative=initial_cumulative,
        flow_type_override=flow_type_override,
    )

    backup_file = _backup_live_file(live_path)

    tag = datetime.now().strftime("%Y%m%d-%H%M%S")
    staged_path = DATA_DIR / f".import-{data_type}-{tag}.xlsx"

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(full_headers)
        for row in live_rows:
            ws.append(row)
        for row in remapped:
            ws.append(row)
        wb.save(staged_path)
        os.replace(staged_path, live_path)
    finally:
        if staged_path.exists() and staged_path != live_path:
            staged_path.unlink(missing_ok=True)

    return {
        "data_type": data_type,
        "label": config["label"],
        "filename": config["filename"],
        "file": str(live_path),
        "backup_file": str(backup_file) if backup_file else None,
        "imported_rows": len(remapped),
        "merged_rows": len(live_rows),
        "mode": mode,
    }


def check_live_has_data(data_type: str) -> bool:
    """Return True if the live file for this data_type already has at least one data row."""
    config = get_data_type_config(data_type)
    live_path = Path(config["path"])
    sheet_name = str(config["sheet_name"])
    return _count_live_rows(live_path, sheet_name) > 0


def export_data_file(data_type: str) -> Path:
    config = get_data_type_config(data_type)
    live_path = Path(config["path"])
    if not live_path.exists():
        raise ValueError(f"No live data file exists yet for {config['label']}.")
    return live_path


def export_all_data_files() -> Path:
    existing_files: list[tuple[str, Path]] = []
    for data_type, config in DATA_TYPE_CONFIG.items():
        live_path = Path(config["path"])
        if live_path.exists():
            existing_files.append((config["filename"], live_path))

    if not existing_files:
        raise ValueError("No data files are available to export yet.")

    tag = datetime.now().strftime("%Y%m%d-%H%M%S")
    zip_path = Path(tempfile.gettempdir()) / f"finledge-data-export-{tag}.zip"

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for filename, live_path in existing_files:
            archive.write(live_path, arcname=filename)

    return zip_path
