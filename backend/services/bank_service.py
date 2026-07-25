import threading
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook, load_workbook

from ..models import BankCategory
from .excel_utils import safe_load_workbook, to_float as _to_float, current_timestamp_us as _current_timestamp
from .path_utils import get_data_dir

DATA_DIR = get_data_dir()
FILE_PATH = DATA_DIR / "bank_transactions.xlsx"
SHEET_NAME = "Bank"
HEADERS = [
    "Date",
    "Category",
    "Amount",
    "Cumulative Amount",
    "Description",
    "Created Timestamp",
    "Last Updated Timestamp",
]
# Derive category list from the Pydantic enum — single source of truth.
BANK_SERVICE_CATEGORIES = [c.value for c in BankCategory]
BANK_INCOME_CATEGORIES = {"interest earned", "income"}

_file_lock = threading.Lock()


def _is_income_category(category: str) -> bool:
    return category.strip().lower() in BANK_INCOME_CATEGORIES


def _ensure_workbook_exists() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    workbook = safe_load_workbook(FILE_PATH, SHEET_NAME, HEADERS)
    modified = False

    # If the file exists but the expected sheet is missing, create/rename it.
    if SHEET_NAME not in workbook.sheetnames:
        modified = True
        if not workbook.sheetnames:
            sheet = workbook.create_sheet(title=SHEET_NAME)
        else:
            candidate = workbook.active
            candidate_first = str(candidate.cell(row=1, column=1).value or "").strip().lower()
            candidate_is_empty = candidate.max_row <= 1 and all(
                candidate.cell(row=1, column=col).value in (None, "") for col in range(1, len(HEADERS) + 1)
            )
            candidate_is_compatible = candidate_first == "date"
            if len(workbook.sheetnames) == 1 and (candidate_is_empty or candidate_is_compatible):
                sheet = candidate
                sheet.title = SHEET_NAME
            else:
                sheet = workbook.create_sheet(title=SHEET_NAME)
    else:
        sheet = workbook[SHEET_NAME]

    # Ensure headers exist without clobbering a first data row.
    first_cell = str(sheet.cell(row=1, column=1).value or "").strip().lower()
    if first_cell != "date":
        modified = True
        sheet.insert_rows(1)
        for idx, header in enumerate(HEADERS, start=1):
            sheet.cell(row=1, column=idx).value = header
    else:
        # Upgrade older files (without Description or timestamp columns) in-place.
        if sheet.max_row >= 1 and sheet.max_column == 4:
            modified = True
            sheet.cell(row=1, column=5).value = "Description"

        # If a header is missing (partial/blank header row), fill it.
        for idx, header in enumerate(HEADERS, start=1):
            if sheet.cell(row=1, column=idx).value in (None, ""):
                modified = True
                sheet.cell(row=1, column=idx).value = header

    if modified:
        workbook.save(FILE_PATH)
    workbook.close()


def append_bank_record(entry_date: Optional[date], category: str, amount: float, description: Optional[str] = None) -> dict:
    with _file_lock:
        _ensure_workbook_exists()

        entry_date = entry_date or date.today()
        timestamp = _current_timestamp()
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[SHEET_NAME]

        previous_cumulative = _to_float(sheet.cell(row=sheet.max_row, column=4).value) if sheet.max_row > 1 else 0.0
        cumulative_amount = previous_cumulative + float(amount)

        sheet.append(
            [
                entry_date.isoformat(),
                category,
                float(amount),
                cumulative_amount,
                (description or "").strip(),
                timestamp,
                timestamp,
            ]
        )
        workbook.save(FILE_PATH)
        workbook.close()

        return {
            "date": entry_date.isoformat(),
            "category": category,
            "amount": float(amount),
            "cumulative_amount": cumulative_amount,
            "description": (description or "").strip() or None,
            "created_timestamp": timestamp,
            "last_updated_timestamp": timestamp,
            "file": str(FILE_PATH),
        }


def read_bank_records() -> list[dict]:
    with _file_lock:
        _ensure_workbook_exists()

        workbook = load_workbook(FILE_PATH, data_only=True)
        sheet = workbook[SHEET_NAME]

        records: list[dict] = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(value is None for value in row):
                continue
            records.append(
                {
                    # Stable ID that matches the Excel row index (header excluded).
                    "id": row_idx - 1,
                    "date": str(row[0] or ""),
                    "category": str(row[1] or ""),
                    "amount": _to_float(row[2]),
                    "cumulative_amount": _to_float(row[3]),
                    "description": str(row[4] or "") if len(row) > 4 else "",
                    "created_timestamp": str(row[5] or "") if len(row) > 5 else "",
                    "last_updated_timestamp": str(row[6] or row[5] or "") if len(row) > 6 else str(row[5] or ""),
                }
            )

        workbook.close()
        return records


def summarize_bank_records(records: list[dict]) -> dict:
    total_income = 0.0
    total_expenses = 0.0
    category_totals = {category: 0.0 for category in BANK_SERVICE_CATEGORIES}

    for record in records:
        category = str(record.get("category") or "").strip()
        amount = _to_float(record.get("amount"))
        if _is_income_category(category):
            total_income += amount
            if category in category_totals:
                category_totals[category] += amount
            elif category:
                category_totals[category] = category_totals.get(category, 0.0) + amount
        else:
            total_expenses += abs(amount)
            if category:
                category_totals[category] = category_totals.get(category, 0.0) + abs(amount)

    net_balance = total_income - total_expenses

    return {
        "total_income": total_income,
        "total_expenses": total_expenses,
        "net_balance": net_balance,
        "category_totals": category_totals,
    }


def _recompute_bank_sheet(sheet) -> None:
    cumulative = 0.0
    for row_idx in range(2, sheet.max_row + 1):
        amount = _to_float(sheet.cell(row=row_idx, column=3).value)
        cumulative += amount
        sheet.cell(row=row_idx, column=4).value = cumulative


def delete_bank_record(record_id: int) -> dict:
    """
    Delete a record by its 1-based index in the data list (header excluded),
    then recompute cumulative amounts so the Excel stays consistent.
    """
    with _file_lock:
        _ensure_workbook_exists()

        if record_id <= 0:
            raise ValueError("record_id must be a positive integer.")

        workbook = load_workbook(FILE_PATH)
        sheet = workbook[SHEET_NAME]

        excel_row = record_id + 1  # +1 for header row
        if excel_row < 2 or excel_row > sheet.max_row:
            raise ValueError("record_id is out of range.")

        sheet.delete_rows(excel_row, 1)
        _recompute_bank_sheet(sheet)
        workbook.save(FILE_PATH)
        workbook.close()

        return {"deleted_id": int(record_id)}


def update_bank_record(
    record_id: int,
    entry_date: Optional[date],
    category: str,
    amount: float,
    description: Optional[str] = None,
) -> dict:
    """
    Update an existing record (1-based, header excluded) and recompute cumulative amounts.
    """
    with _file_lock:
        _ensure_workbook_exists()

        if record_id <= 0:
            raise ValueError("record_id must be a positive integer.")

        entry_date = entry_date or date.today()
        timestamp = _current_timestamp()

        workbook = load_workbook(FILE_PATH)
        sheet = workbook[SHEET_NAME]

        excel_row = record_id + 1  # +1 for header row
        if excel_row < 2 or excel_row > sheet.max_row:
            raise ValueError("record_id is out of range.")

        sheet.cell(row=excel_row, column=1).value = entry_date.isoformat()
        sheet.cell(row=excel_row, column=2).value = category
        sheet.cell(row=excel_row, column=3).value = float(amount)
        sheet.cell(row=excel_row, column=5).value = (description or "").strip()
        # Created Timestamp remains immutable; only Last Updated Timestamp changes.
        if not sheet.cell(row=excel_row, column=6).value:
            sheet.cell(row=excel_row, column=6).value = timestamp
        sheet.cell(row=excel_row, column=7).value = timestamp

        _recompute_bank_sheet(sheet)
        workbook.save(FILE_PATH)
        workbook.close()

        return {
            "updated_id": int(record_id),
            "date": entry_date.isoformat(),
            "category": category,
            "amount": float(amount),
            "description": (description or "").strip() or None,
            "created_timestamp": str(sheet.cell(row=excel_row, column=6).value or timestamp),
            "last_updated_timestamp": timestamp,
            "file": str(FILE_PATH),
        }

