from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Optional

from openpyxl import Workbook, load_workbook

from .path_utils import get_data_dir

DATA_DIR = get_data_dir()
FILE_PATH = DATA_DIR / "share_transactions.xlsx"
SHEET_NAME = "Share"
HEADERS = [
    "Date",
    "Share Name",
    "Category",
    "Per Unit Price",
    "ASBA Charge",
    "Allotted",
    "Buy/Sell",
    "Total Amount",
    "Profit/Loss",
    "Cumulative Profit",
]


def _to_float(value: object) -> float:
    try:
        return float(value or 0.0)
    except (TypeError, ValueError):
        return 0.0


def _to_int(value: object) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def _ensure_workbook_exists() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    if not FILE_PATH.exists():
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME
        sheet.append(HEADERS)
        workbook.save(FILE_PATH)
        return

    workbook = load_workbook(FILE_PATH)

    # If the file exists but the expected sheet is missing, create/rename it.
    if SHEET_NAME not in workbook.sheetnames:
        if not workbook.sheetnames:
            sheet = workbook.create_sheet(title=SHEET_NAME)
        else:
            candidate = workbook.active
            candidate_first = str(candidate.cell(row=1, column=1).value or "").strip().lower()
            candidate_is_empty = candidate.max_row <= 1 and all(
                candidate.cell(row=1, column=col).value in (None, "") for col in range(1, len(HEADERS) + 1)
            )
            candidate_is_compatible = candidate_first == "date"
            if len(workbook.sheetnames) == 1 and (candidate_is_empty or candidate_is_compatible):
                sheet = candidate
                sheet.title = SHEET_NAME
            else:
                sheet = workbook.create_sheet(title=SHEET_NAME)
    else:
        sheet = workbook[SHEET_NAME]

    # Ensure headers exist without clobbering a first data row.
    first_cell = str(sheet.cell(row=1, column=1).value or "").strip().lower()
    if first_cell != "date":
        sheet.insert_rows(1)
        for idx, header in enumerate(HEADERS, start=1):
            sheet.cell(row=1, column=idx).value = header
    else:
        for idx, header in enumerate(HEADERS, start=1):
            if sheet.cell(row=1, column=idx).value in (None, ""):
                sheet.cell(row=1, column=idx).value = header

    workbook.save(FILE_PATH)


def _build_open_lots(sheet, share_name: str) -> list[dict]:
    lots: list[dict] = []
    target = share_name.strip().lower()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_share_name = str(row[1] or "").strip().lower()
        if row_share_name != target:
            continue

        category = str(row[2] or "").strip().lower()
        qty = _to_int(row[5])
        price = _to_float(row[3])
        asba_charge = _to_float(row[4])
        buy_sell = str(row[6] or "").strip().lower()

        if qty <= 0:
            continue

        if category in {"ipo", "buy"} or (category == "dividend" and buy_sell == "bonus"):
            lots.append(
                {
                    "qty": qty,
                    "price": price,
                    # ASBA is part of IPO investment. Secondary buys have 0 ASBA.
                    "asba": asba_charge if category == "ipo" else 0.0,
                }
            )
            continue

        if category == "sell":
            remaining = qty
            for lot in lots:
                if remaining <= 0:
                    break
                qty_before = lot["qty"]
                consumed = min(qty_before, remaining)
                # Consume ASBA proportionally from the remaining lot.
                if qty_before > 0 and lot.get("asba", 0.0) != 0.0:
                    lot["asba"] -= lot["asba"] * (consumed / qty_before)
                lot["qty"] = qty_before - consumed
                remaining -= consumed
            lots = [lot for lot in lots if lot["qty"] > 0]

    return lots


def _cost_basis_for_sell(lots: list[dict], sell_qty: int) -> float:
    remaining = sell_qty
    cost_basis = 0.0

    for lot in lots:
        if remaining <= 0:
            break
        qty_available = lot["qty"]
        consumed = min(qty_available, remaining)
        cost_basis += consumed * lot["price"]

        # Include remaining IPO ASBA as part of cost basis, proportional to shares sold.
        asba_remaining = float(lot.get("asba", 0.0) or 0.0)
        if qty_available > 0 and asba_remaining != 0.0:
            cost_basis += asba_remaining * (consumed / qty_available)
        remaining -= consumed

    if remaining > 0:
        raise ValueError("Not enough available quantity to sell for this share.")

    return cost_basis


def _consume_lots(lots: list[dict], sell_qty: int) -> float:
    remaining = sell_qty
    cost_basis = 0.0

    for lot in lots:
        if remaining <= 0:
            break
        qty_before = lot["qty"]
        consumed = min(qty_before, remaining)
        cost_basis += consumed * lot["price"]

        asba_remaining = float(lot.get("asba", 0.0) or 0.0)
        if qty_before > 0 and asba_remaining != 0.0:
            consumed_asba = asba_remaining * (consumed / qty_before)
            cost_basis += consumed_asba
            lot["asba"] = asba_remaining - consumed_asba

        lot["qty"] = qty_before - consumed
        remaining -= consumed

    if remaining > 0:
        raise ValueError("Not enough available quantity to sell for this share.")

    return cost_basis


def append_share_record(
    entry_date: Optional[date],
    share_name: str,
    category: str,
    per_unit_price: Decimal,
    allotted: int,
    buy_sell: str,
) -> dict:
    _ensure_workbook_exists()

    entry_date = entry_date or date.today()
    share_name = share_name.strip().upper()
    category = category.strip().lower()
    if category not in {"ipo", "buy", "sell", "dividend"}:
        raise ValueError("Category must be one of: ipo, buy, sell, dividend.")

    if category == "dividend":
        buy_sell = buy_sell.strip().lower()
        if buy_sell not in {"cash", "bonus"}:
            raise ValueError("Dividend type must be 'cash' or 'bonus'.")
        if buy_sell == "cash" and allotted != 0:
            raise ValueError("Cash dividend must have 0 allotted shares.")
        if buy_sell == "bonus" and allotted <= 0:
            raise ValueError("Bonus dividend must have greater than 0 allotted shares.")
    else:
        if allotted < 0:
            raise ValueError("Allotted cannot be negative.")
        if category != "ipo" and allotted <= 0:
            raise ValueError("Allotted must be greater than 0 for buy/sell entries.")

    if per_unit_price < 0:
        raise ValueError("Per unit price cannot be negative.")

    workbook = load_workbook(FILE_PATH)
    sheet = workbook[SHEET_NAME]

    # ASBA charge rules:
    # - ipo: 5
    # - secondary/buy/sell: 0
    asba_charge_dec = Decimal("5") if category == "ipo" else Decimal("0")

    # Keep currency math in Decimal to avoid float rounding surprises.
    try:
        unit_price = per_unit_price if isinstance(per_unit_price, Decimal) else Decimal(str(per_unit_price))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError("Per unit price must be a valid number.") from exc

    if category == "dividend":
        if buy_sell == "cash":
            total_amount_dec = unit_price
            profit_loss_dec = total_amount_dec
        else:
            total_amount_dec = Decimal("0")
            profit_loss_dec = Decimal("0")
    else:
        share_value_dec = unit_price * Decimal(int(allotted))
        total_amount_dec = share_value_dec + asba_charge_dec
        profit_loss_dec = Decimal("0")
        if category == "sell":
            open_lots = _build_open_lots(sheet, share_name)
            buy_cost = _cost_basis_for_sell(open_lots, int(allotted))
            profit_loss_dec = total_amount_dec - Decimal(str(buy_cost))

    previous_cumulative = _to_float(sheet.cell(row=sheet.max_row, column=10).value) if sheet.max_row > 1 else 0.0
    cumulative_profit = previous_cumulative + float(profit_loss_dec)

    # Debug: helps confirm the exact value we store vs what the client sent.
    print(
        f"[share] saving share_name={share_name} unit_price={unit_price} "
        f"allotted={allotted} category={category} asba_charge={asba_charge_dec} total={total_amount_dec}"
    )

    sheet.append(
        [
            entry_date.isoformat(),
            share_name,
            category,
            str(unit_price),
            float(asba_charge_dec),
            int(allotted),
            buy_sell,
            str(total_amount_dec),
            str(profit_loss_dec),
            cumulative_profit,
        ]
    )
    workbook.save(FILE_PATH)

    return {
        "date": entry_date.isoformat(),
        "share_name": share_name,
        "category": category,
        "per_unit_price": str(unit_price),
        "asba_charge": float(asba_charge_dec),
        "allotted": int(allotted),
        "buy_sell": buy_sell,
        "total_amount": str(total_amount_dec),
        "profit_loss": str(profit_loss_dec),
        "cumulative_profit": cumulative_profit,
        "file": str(FILE_PATH),
    }


def read_share_records() -> list[dict]:
    _ensure_workbook_exists()

    workbook = load_workbook(FILE_PATH, data_only=True)
    sheet = workbook[SHEET_NAME]

    records: list[dict] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(value is None for value in row):
            continue
        records.append(
            {
                # Stable ID that matches the Excel row index (header excluded).
                "id": row_idx - 1,
                "date": str(row[0] or ""),
                "share_name": str(row[1] or ""),
                "category": str(row[2] or ""),
                "per_unit_price": _to_float(row[3]),
                "asba_charge": _to_float(row[4]),
                "allotted": _to_int(row[5]),
                "buy_sell": str(row[6] or ""),
                "total_amount": _to_float(row[7]),
                "profit_loss": _to_float(row[8]),
                "cumulative_profit": _to_float(row[9]),
            }
        )

    return records


def summarize_share_records(records: list[dict]) -> dict:
    total_ipo_investment = 0.0
    total_buy_amount = 0.0
    total_sell_amount = 0.0
    total_profit = 0.0
    total_dividend = 0.0

    for record in records:
        category = str(record.get("category") or "").strip().lower()
        buy_sell = str(record.get("buy_sell") or "").strip().lower()
        total_amount = _to_float(record.get("total_amount"))
        profit_loss = _to_float(record.get("profit_loss"))

        if category == "ipo":
            total_ipo_investment += total_amount
        elif category == "buy":
            total_buy_amount += total_amount
        elif category == "sell":
            total_sell_amount += total_amount
            total_profit += profit_loss
        elif category == "dividend" and buy_sell == "cash":
            total_dividend += total_amount

    overall_investment = total_ipo_investment + total_buy_amount
    overall_profit_loss = total_profit + total_dividend - overall_investment

    return {
        "total_ipo_investment": total_ipo_investment,
        "total_buy_amount": total_buy_amount,
        "overall_investment": overall_investment,
        "total_sell_amount": total_sell_amount,
        "total_dividend": total_dividend,
        "total_profit": total_profit,
        "overall_profit_loss": overall_profit_loss,
    }


def _recompute_sheet(sheet) -> None:
    cumulative_profit = 0.0
    lots_by_share: dict[str, list[dict]] = {}

    for row_idx in range(2, sheet.max_row + 1):
        share_name = str(sheet.cell(row=row_idx, column=2).value or "").strip()
        category = str(sheet.cell(row=row_idx, column=3).value or "").strip().lower()
        per_unit_price = _to_float(sheet.cell(row=row_idx, column=4).value)
        allotted = _to_int(sheet.cell(row=row_idx, column=6).value)
        buy_sell = str(sheet.cell(row=row_idx, column=7).value or "")

        asba_charge = 5.0 if category == "ipo" else 0.0
        profit_loss = 0.0
        share_key = share_name.lower()
        
        if category == "dividend":
            if buy_sell == "cash":
                total_amount = float(per_unit_price)
                profit_loss = total_amount
            else:
                total_amount = 0.0
        else:
            total_amount = (float(per_unit_price) * int(allotted)) + float(asba_charge)

        if category in {"ipo", "buy"} or (category == "dividend" and buy_sell == "bonus"):
            if share_key:
                lots_by_share.setdefault(share_key, []).append(
                    {"qty": allotted, "price": per_unit_price if category != "dividend" else 0.0, "asba": asba_charge if category == "ipo" else 0.0}
                )
        elif category == "sell":
            lots = lots_by_share.get(share_key, [])
            if allotted > 0:
                cost_basis = _consume_lots(lots, allotted)
                profit_loss = total_amount - cost_basis

        cumulative_profit += profit_loss

        sheet.cell(row=row_idx, column=5).value = asba_charge
        sheet.cell(row=row_idx, column=7).value = buy_sell
        sheet.cell(row=row_idx, column=8).value = total_amount
        sheet.cell(row=row_idx, column=9).value = profit_loss
        sheet.cell(row=row_idx, column=10).value = cumulative_profit


def update_share_allotment(share_name: str, allotted: int) -> dict:
    _ensure_workbook_exists()

    workbook = load_workbook(FILE_PATH)
    sheet = workbook[SHEET_NAME]

    target = share_name.strip().lower()
    target_row = None
    found_any = False

    for row_idx in range(sheet.max_row, 1, -1):
        row_share_name = str(sheet.cell(row=row_idx, column=2).value or "").strip().lower()
        category = str(sheet.cell(row=row_idx, column=3).value or "").strip().lower()
        if row_share_name == target:
            found_any = True
        if row_share_name == target and category == "ipo":
            target_row = row_idx
            break

    if not target_row:
        if found_any:
            raise ValueError("Only IPO entries can be updated from this form.")
        raise ValueError("No IPO entry found for the provided share name.")

    previous_allotted = _to_int(sheet.cell(row=target_row, column=6).value)
    updated_date = str(sheet.cell(row=target_row, column=1).value or "")
    normalized_share_name = str(sheet.cell(row=target_row, column=2).value or share_name).strip().upper()

    sheet.cell(row=target_row, column=6).value = int(allotted)
    _recompute_sheet(sheet)
    workbook.save(FILE_PATH)

    return {
        "record_id": int(target_row - 1),
        "date": updated_date,
        "share_name": normalized_share_name,
        "previous_allotted": previous_allotted,
        "allotted": int(allotted),
    }


def delete_share_record(record_id: int) -> dict:
    """
    Delete a record by its 1-based index in the data list (header excluded),
    then recompute ASBA/total/profit/cumulative so the Excel stays consistent.
    """
    _ensure_workbook_exists()

    if record_id <= 0:
        raise ValueError("record_id must be a positive integer.")

    workbook = load_workbook(FILE_PATH)
    sheet = workbook[SHEET_NAME]

    excel_row = record_id + 1  # +1 for header row
    if excel_row < 2 or excel_row > sheet.max_row:
        raise ValueError("record_id is out of range.")

    sheet.delete_rows(excel_row, 1)
    _recompute_sheet(sheet)
    workbook.save(FILE_PATH)

    return {"deleted_id": int(record_id)}


def update_share_record(
    record_id: int,
    entry_date: Optional[date],
    share_name: str,
    category: str,
    per_unit_price: Decimal,
    allotted: int,
    buy_sell: str,
) -> dict:
    """
    Update an existing record (1-based, header excluded) and recompute totals/profit/cumulative.

    Note: recomputation keeps Excel consistent after edits.
    """
    _ensure_workbook_exists()

    if record_id <= 0:
        raise ValueError("record_id must be a positive integer.")

    entry_date = entry_date or date.today()
    share_name = share_name.strip().upper()
    category = category.strip().lower()
    if category not in {"ipo", "buy", "sell", "dividend"}:
        raise ValueError("Category must be one of: ipo, buy, sell, dividend.")

    if category == "dividend":
        buy_sell = (buy_sell or "").strip().lower()
        if buy_sell not in {"cash", "bonus"}:
            raise ValueError("Dividend type must be 'cash' or 'bonus'.")
        if buy_sell == "cash" and allotted != 0:
            raise ValueError("Cash dividend must have 0 allotted shares.")
        if buy_sell == "bonus" and allotted <= 0:
            raise ValueError("Bonus dividend must have greater than 0 allotted shares.")
    else:
        if allotted < 0:
            raise ValueError("Allotted cannot be negative.")
        if category != "ipo" and allotted <= 0:
            raise ValueError("Allotted must be greater than 0 for buy/sell entries.")

    if per_unit_price < 0:
        raise ValueError("Per unit price cannot be negative.")

    try:
        unit_price = per_unit_price if isinstance(per_unit_price, Decimal) else Decimal(str(per_unit_price))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError("Per unit price must be a valid number.") from exc

    workbook = load_workbook(FILE_PATH)
    sheet = workbook[SHEET_NAME]

    excel_row = record_id + 1  # +1 for header row
    if excel_row < 2 or excel_row > sheet.max_row:
        raise ValueError("record_id is out of range.")

    sheet.cell(row=excel_row, column=1).value = entry_date.isoformat()
    sheet.cell(row=excel_row, column=2).value = share_name
    sheet.cell(row=excel_row, column=3).value = category
    # Store as string to avoid client-side floating drift and keep what the user entered.
    sheet.cell(row=excel_row, column=4).value = str(unit_price)
    sheet.cell(row=excel_row, column=6).value = int(allotted)
    sheet.cell(row=excel_row, column=7).value = str(buy_sell or category).strip().lower()

    _recompute_sheet(sheet)
    workbook.save(FILE_PATH)

    return {
        "updated_id": int(record_id),
        "date": entry_date.isoformat(),
        "share_name": share_name,
        "category": category,
        "per_unit_price": str(unit_price),
        "allotted": int(allotted),
        "buy_sell": str(buy_sell or category).strip().lower(),
        "file": str(FILE_PATH),
    }
