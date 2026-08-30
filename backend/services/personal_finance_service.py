import threading
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook, load_workbook

from .excel_utils import safe_load_workbook, to_float as _to_float, current_timestamp as _current_timestamp
from .path_utils import get_data_dir

DATA_DIR = get_data_dir()
BANK_FLOW_FILE_PATH = DATA_DIR / "personal_finance_bank_flow.xlsx"
CASH_FLOW_FILE_PATH = DATA_DIR / "personal_finance_cash_flow.xlsx"
FILE_PATH = BANK_FLOW_FILE_PATH
SHEET_NAME = "Personal Finance"
HEADERS = [
    "Date",
    "Flow Type",
    "Direction",
    "Category",
    "Amount",
    "Signed Amount",
    "Description",
    "Source",
    "Created Timestamp",
    "Last Updated Timestamp",
    "Source Ref",
    "Updated Device",
]

_file_lock = threading.Lock()


def _normalize_flow_type(flow_type: str) -> str:
    normalized = str(flow_type or "").strip().lower()
    if normalized not in {"bank", "cash"}:
        raise ValueError("flow_type must be 'bank' or 'cash'.")
    return normalized


def _flow_file_path(flow_type: str):
    return BANK_FLOW_FILE_PATH if _normalize_flow_type(flow_type) == "bank" else CASH_FLOW_FILE_PATH


def _flow_sheet_name(flow_type: str) -> str:
    return "Bank Flow" if _normalize_flow_type(flow_type) == "bank" else "Cash Flow"


def _ensure_workbook_exists(flow_type: str) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    file_path = _flow_file_path(flow_type)
    sheet_name = _flow_sheet_name(flow_type)

    workbook = safe_load_workbook(file_path, sheet_name, HEADERS)
    modified = False

    created_sheet = False
    if sheet_name not in workbook.sheetnames:
        modified = True
        sheet = workbook.active if workbook.sheetnames else workbook.create_sheet(title=sheet_name)
        if sheet.max_row <= 1:
            sheet.title = sheet_name
        else:
            sheet = workbook.create_sheet(title=sheet_name)
        created_sheet = True
    else:
        sheet = workbook[sheet_name]

    first_cell = str(sheet.cell(row=1, column=1).value or "").strip().lower()
    if first_cell != "date" and not created_sheet:
        modified = True
        sheet.insert_rows(1)

    for idx, header in enumerate(HEADERS, start=1):
        if sheet.cell(row=1, column=idx).value in (None, ""):
            modified = True
            sheet.cell(row=1, column=idx).value = header

    if modified:
        workbook.save(file_path)
    workbook.close()


def _signed_amount(direction: str, amount: Decimal | float) -> float:
    value = float(amount)
    return value if direction == "income" else -value


def append_personal_finance_record(
    entry_date: Optional[date],
    flow_type: str,
    direction: str,
    category: str,
    amount: Decimal | float,
    description: Optional[str] = None,
    source: str = "manual",
    source_ref: Optional[str] = None,
) -> dict:
    with _file_lock:
        flow_type = _normalize_flow_type(flow_type)
        _ensure_workbook_exists(flow_type)

        entry_date = entry_date or date.today()
        timestamp = _current_timestamp()
        signed_amount = _signed_amount(direction, amount)
        amount_float = float(amount)
        description_value = (description or "").strip()

        file_path = _flow_file_path(flow_type)
        sheet_name = _flow_sheet_name(flow_type)
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        sheet.append(
            [
                entry_date.isoformat(),
                flow_type,
                direction,
                category,
                amount_float,
                signed_amount,
                description_value,
                source,
                timestamp,
                timestamp,
                source_ref or "",
                "desktop",
            ]
        )
        workbook.save(file_path)
        workbook.close()

        return {
            "date": entry_date.isoformat(),
            "flow_type": flow_type,
            "direction": direction,
            "category": category,
            "amount": amount_float,
            "signed_amount": signed_amount,
            "description": description_value or None,
            "source": source,
            "timestamp": timestamp,
            "created_timestamp": timestamp,
            "last_updated_timestamp": timestamp,
            "source_ref": source_ref or "",
            "updated_device": "desktop",
            "file": str(file_path),
        }


def _read_personal_finance_records_for_flow(flow_type: str) -> list[dict]:
    with _file_lock:
        flow_type = _normalize_flow_type(flow_type)
        _ensure_workbook_exists(flow_type)

        file_path = _flow_file_path(flow_type)
        sheet_name = _flow_sheet_name(flow_type)
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook[sheet_name]

        records: list[dict] = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(value is None for value in row):
                continue

            amount = _to_float(row[4] if len(row) > 4 else 0)
            signed_amount = _to_float(row[5] if len(row) > 5 else 0)
            direction = str(row[2] or "").strip().lower()
            if not signed_amount and amount:
                signed_amount = _signed_amount(direction, amount)

            source = str(row[7] or "manual") if len(row) > 7 else "manual"
            # Share sync rows from the first implementation were physical copies.
            # Bank Flow is now derived live from the Share Portfolio workbook, so
            # omit those old copies to keep storage separate and prevent duplicates.
            if source == "share-sync":
                continue

            records.append(
                {
                    "id": row_idx - 1,
                    "display_id": f"{flow_type[:1].upper()}-{row_idx - 1}",
                    "date": str(row[0] or ""),
                    "flow_type": flow_type,
                    "direction": direction,
                    "category": str(row[3] or ""),
                    "amount": amount,
                    "signed_amount": signed_amount,
                    "description": str(row[6] or "") if len(row) > 6 else "",
                    "source": source,
                    "timestamp": str(row[9] or row[8] or "") if len(row) > 9 else str(row[8] or "") if len(row) > 8 else "",
                    "created_timestamp": str(row[8] or "") if len(row) > 8 else "",
                    "last_updated_timestamp": str(row[9] or row[8] or "") if len(row) > 9 else str(row[8] or "") if len(row) > 8 else "",
                    "source_ref": str(row[10] or "") if len(row) > 10 else str(row[9] or "") if len(row) > 9 else "",
                    "updated_device": str(row[11] or "legacy") if len(row) > 11 else "legacy",
                }
            )

        workbook.close()
        return records


def _share_sync_mapping(record: dict) -> tuple[str, str] | None:
    category = str(record.get("category") or "").strip().lower()
    buy_sell = str(record.get("buy_sell") or "").strip().lower()

    if category in {"ipo", "buy"}:
        return "expense", "Investment Expense"
    if category == "sip" and buy_sell in {"redeem", "redeemed"}:
        return "income", "Investment Income"
    if category == "sip":
        return "expense", "Investment Expense"
    if category == "sell":
        return "income", "Investment Income"
    if category == "dividend" and buy_sell == "cash":
        return "income", "Investment Income"
    return None


def _build_share_sync_records() -> list[dict]:
    # Import here to avoid a service import cycle during app start-up.
    from backend.services.share_service import read_share_records

    records: list[dict] = []
    for source_record in read_share_records():
        mapping = _share_sync_mapping(source_record)
        amount = abs(_to_float(source_record.get("total_amount")))
        if not mapping or amount <= 0:
            continue

        direction, category = mapping
        share_category = str(source_record.get("category") or "").strip().lower()
        buy_sell = str(source_record.get("buy_sell") or "").strip().lower()
        event_labels = {
            ("ipo", "ipo"): "IPO",
            ("buy", "buy"): "Secondary buy",
            ("sell", "sell"): "Share sell",
            ("sip", "installment"): "SIP installment",
            ("sip", "redeem"): "SIP redeem",
            ("sip", "redeemed"): "SIP redeem",
            ("dividend", "cash"): "Cash dividend",
        }
        source_id = int(source_record.get("id") or 0)
        records.append(
            {
                "id": f"share-{source_id}",
                "display_id": f"S-{source_id}",
                "date": str(source_record.get("date") or ""),
                "flow_type": "bank",
                "direction": direction,
                "category": category,
                "amount": amount,
                "signed_amount": _signed_amount(direction, amount),
                "description": f"{event_labels.get((share_category, buy_sell), share_category.title())}: {str(source_record.get('share_name') or '').strip().upper()}",
                "source": "share-sync",
                "timestamp": str(source_record.get("last_updated_timestamp") or source_record.get("timestamp") or ""),
                "source_ref": str(source_record.get("sync_ref") or f"share:{source_id}"),
            }
        )
    return records


def _build_bank_services_sync_records() -> list[dict]:
    # Bank Services remains the source of truth. Its costs are displayed in
    # Bank Flow as Service Cost and its interest as Interest Earned.
    from backend.services.bank_service import read_bank_records

    records: list[dict] = []
    for source_record in read_bank_records():
        original_category = str(source_record.get("category") or "").strip()
        amount = abs(_to_float(source_record.get("amount")))
        if amount <= 0:
            continue

        is_interest = original_category.lower() in {"interest earned", "income"}
        direction = "income" if is_interest else "expense"
        category = "Interest Earned" if is_interest else "Service Cost"
        source_id = int(source_record.get("id") or 0)
        description = str(source_record.get("description") or "").strip()
        label = original_category or category
        records.append(
            {
                "id": f"bank-services-{source_id}",
                "display_id": f"BS-{source_id}",
                "date": str(source_record.get("date") or ""),
                "flow_type": "bank",
                "direction": direction,
                "category": category,
                "amount": amount,
                "signed_amount": _signed_amount(direction, amount),
                "description": f"{label}{': ' + description if description else ''}",
                "source": "bank-services-sync",
                "timestamp": str(source_record.get("timestamp") or ""),
                "source_ref": f"bank-services:{source_id}",
            }
        )
    return records


def read_personal_finance_records(flow_type: Optional[str] = None) -> list[dict]:
    normalized_flow = str(flow_type or "combined").strip().lower()
    bank_records = (
        _read_personal_finance_records_for_flow("bank")
        + _build_share_sync_records()
        + _build_bank_services_sync_records()
    )
    if normalized_flow == "bank":
        records = bank_records
    elif normalized_flow == "cash":
        records = _read_personal_finance_records_for_flow("cash")
    else:
        records = bank_records + _read_personal_finance_records_for_flow("cash")
    return sorted(
        records,
        key=lambda record: (
            str(record.get("timestamp") or ""),
            str(record.get("date") or ""),
            str(record.get("display_id") or ""),
        ),
    )


def _empty_flow_summary() -> dict:
    return {
        "income": 0.0,
        "expenses": 0.0,
        "investment_expense": 0.0,
        "investment_income": 0.0,
        "interest_earned": 0.0,
        "service_cost": 0.0,
        "total_income": 0.0,
        "total_expenses": 0.0,
        "net": 0.0,
        "income_breakdown": {},
        "expense_breakdown": {},
    }


def summarize_personal_finance_records(records: list[dict]) -> dict:
    flow_summaries = {
        "bank": _empty_flow_summary(),
        "cash": _empty_flow_summary(),
    }

    for record in records:
        flow_type = str(record.get("flow_type") or "").strip().lower()
        if flow_type not in flow_summaries:
            continue

        direction = str(record.get("direction") or "").strip().lower()
        category = str(record.get("category") or "").strip() or "Uncategorized"
        amount = abs(_to_float(record.get("amount")))
        summary = flow_summaries[flow_type]

        source = str(record.get("source") or "manual").strip().lower()
        category_key = category.lower()

        if direction == "income":
            summary["total_income"] += amount
            summary["income_breakdown"][category] = summary["income_breakdown"].get(category, 0.0) + amount
            if flow_type == "bank" and source == "bank-services-sync":
                summary["interest_earned"] += amount
            elif flow_type == "bank" and category_key == "investment income":
                summary["investment_income"] += amount
            else:
                summary["income"] += amount
        else:
            summary["total_expenses"] += amount
            summary["expense_breakdown"][category] = summary["expense_breakdown"].get(category, 0.0) + amount
            if flow_type == "bank" and source == "bank-services-sync":
                summary["service_cost"] += amount
            elif flow_type == "bank" and category_key in {"investment expense", "investment", "sip", "share market"}:
                summary["investment_expense"] += amount
            else:
                summary["expenses"] += amount

    for summary in flow_summaries.values():
        summary["net"] = summary["total_income"] - summary["total_expenses"]

    combined = {
        "overall_income": flow_summaries["bank"]["total_income"] + flow_summaries["cash"]["total_income"],
        "overall_expenses": flow_summaries["bank"]["total_expenses"] + flow_summaries["cash"]["total_expenses"],
        "bank": flow_summaries["bank"],
        "cash": flow_summaries["cash"],
    }
    combined["overall_net"] = combined["overall_income"] - combined["overall_expenses"]

    return {
        "bank": flow_summaries["bank"],
        "cash": flow_summaries["cash"],
        "combined": combined,
    }


def delete_personal_finance_record(record_id: int, flow_type: str) -> dict:
    with _file_lock:
        flow_type = _normalize_flow_type(flow_type)
        _ensure_workbook_exists(flow_type)

        if record_id <= 0:
            raise ValueError("record_id must be a positive integer.")

        file_path = _flow_file_path(flow_type)
        sheet_name = _flow_sheet_name(flow_type)
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        excel_row = record_id + 1
        if excel_row < 2 or excel_row > sheet.max_row:
            raise ValueError("record_id is out of range.")

        source = str(sheet.cell(row=excel_row, column=8).value or "manual")
        if source != "manual":
            raise ValueError("Synced Personal Expenses records are read-only.")

        sheet.delete_rows(excel_row, 1)
        workbook.save(file_path)
        workbook.close()
        return {"deleted_id": int(record_id), "flow_type": flow_type}


def update_personal_finance_record(
    record_id: int,
    record_flow_type: str,
    entry_date: Optional[date],
    flow_type: str,
    direction: str,
    category: str,
    amount: Decimal | float,
    description: Optional[str] = None,
    source: str = "manual",
) -> dict:
    with _file_lock:
        record_flow_type = _normalize_flow_type(record_flow_type)
        flow_type = _normalize_flow_type(flow_type)
        _ensure_workbook_exists(record_flow_type)

        if record_id <= 0:
            raise ValueError("record_id must be a positive integer.")

        file_path = _flow_file_path(record_flow_type)
        sheet_name = _flow_sheet_name(record_flow_type)
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        excel_row = record_id + 1
        if excel_row < 2 or excel_row > sheet.max_row:
            raise ValueError("record_id is out of range.")

        existing_source = str(sheet.cell(row=excel_row, column=8).value or "manual")
        if existing_source != "manual":
            raise ValueError("Synced Personal Expenses records are read-only.")

        if record_flow_type != flow_type:
            sheet.delete_rows(excel_row, 1)
            workbook.save(file_path)
            workbook.close()
            return append_personal_finance_record(
                entry_date=entry_date,
                flow_type=flow_type,
                direction=direction,
                category=category,
                amount=amount,
                description=description,
                source=source,
            )

        entry_date = entry_date or date.today()
        timestamp = _current_timestamp()
        signed_amount = _signed_amount(direction, amount)
        amount_float = float(amount)
        description_value = (description or "").strip()

        sheet.cell(row=excel_row, column=1).value = entry_date.isoformat()
        sheet.cell(row=excel_row, column=2).value = flow_type
        sheet.cell(row=excel_row, column=3).value = direction
        sheet.cell(row=excel_row, column=4).value = category
        sheet.cell(row=excel_row, column=5).value = amount_float
        sheet.cell(row=excel_row, column=6).value = signed_amount
        sheet.cell(row=excel_row, column=7).value = description_value
        sheet.cell(row=excel_row, column=8).value = source
        if not sheet.cell(row=excel_row, column=9).value:
            sheet.cell(row=excel_row, column=9).value = timestamp
        sheet.cell(row=excel_row, column=10).value = timestamp
        sheet.cell(row=excel_row, column=11).value = ""
        sheet.cell(row=excel_row, column=12).value = "desktop"

        workbook.save(file_path)
        workbook.close()
        return {
            "updated_id": int(record_id),
            "date": entry_date.isoformat(),
            "flow_type": flow_type,
            "direction": direction,
            "category": category,
            "amount": amount_float,
            "signed_amount": signed_amount,
            "description": description_value or None,
            "source": source,
            "timestamp": timestamp,
            "created_timestamp": str(sheet.cell(row=excel_row, column=9).value or timestamp),
            "last_updated_timestamp": timestamp,
            "source_ref": "",
            "updated_device": "desktop",
            "file": str(file_path),
        }


def upsert_share_sync_record(
    *,
    source_ref: str,
    entry_date: Optional[date] | str,
    direction: str,
    category: str,
    amount: Decimal | float,
    description: Optional[str] = None,
) -> dict:
    with _file_lock:
        if not source_ref:
            raise ValueError("source_ref is required for share sync.")

        flow_type = "bank"
        _ensure_workbook_exists(flow_type)
        file_path = _flow_file_path(flow_type)
        sheet_name = _flow_sheet_name(flow_type)
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        if isinstance(entry_date, date):
            date_value = entry_date.isoformat()
        else:
            date_value = str(entry_date or date.today().isoformat())

        timestamp = _current_timestamp()
        amount_float = abs(float(amount))
        signed_amount = _signed_amount(direction, amount_float)
        description_value = (description or "").strip()

        target_row = None
        for row_idx in range(2, sheet.max_row + 1):
            existing_source = str(sheet.cell(row=row_idx, column=8).value or "").strip()
            existing_ref = str(sheet.cell(row=row_idx, column=11).value or "").strip()
            if existing_source == "share-sync" and existing_ref == source_ref:
                target_row = row_idx
                break

        if target_row is None:
            sheet.append(
                [
                    date_value,
                    flow_type,
                    direction,
                    category,
                    amount_float,
                    signed_amount,
                    description_value,
                    "share-sync",
                    timestamp,
                    timestamp,
                    source_ref,
                    "desktop",
                ]
            )
            record_id = sheet.max_row - 1
        else:
            sheet.cell(row=target_row, column=1).value = date_value
            sheet.cell(row=target_row, column=2).value = flow_type
            sheet.cell(row=target_row, column=3).value = direction
            sheet.cell(row=target_row, column=4).value = category
            sheet.cell(row=target_row, column=5).value = amount_float
            sheet.cell(row=target_row, column=6).value = signed_amount
            sheet.cell(row=target_row, column=7).value = description_value
            sheet.cell(row=target_row, column=8).value = "share-sync"
            if not sheet.cell(row=target_row, column=9).value:
                sheet.cell(row=target_row, column=9).value = timestamp
            sheet.cell(row=target_row, column=10).value = timestamp
            sheet.cell(row=target_row, column=11).value = source_ref
            sheet.cell(row=target_row, column=12).value = "desktop"
            record_id = target_row - 1

        workbook.save(file_path)
        workbook.close()
        return {
            "record_id": int(record_id),
            "date": date_value,
            "flow_type": flow_type,
            "direction": direction,
            "category": category,
            "amount": amount_float,
            "signed_amount": signed_amount,
            "description": description_value or None,
            "source": "share-sync",
            "source_ref": source_ref,
            "timestamp": timestamp,
            "created_timestamp": timestamp,
            "last_updated_timestamp": timestamp,
            "updated_device": "desktop",
            "file": str(file_path),
        }


def delete_share_sync_record(source_ref: str) -> dict:
    with _file_lock:
        if not source_ref:
            return {"deleted": False, "source_ref": ""}

        flow_type = "bank"
        _ensure_workbook_exists(flow_type)
        file_path = _flow_file_path(flow_type)
        sheet_name = _flow_sheet_name(flow_type)
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        for row_idx in range(sheet.max_row, 1, -1):
            existing_source = str(sheet.cell(row=row_idx, column=8).value or "").strip()
            existing_ref = str(sheet.cell(row=row_idx, column=11).value or "").strip()
            if existing_source == "share-sync" and existing_ref == source_ref:
                sheet.delete_rows(row_idx, 1)
                workbook.save(file_path)
                workbook.close()
                return {"deleted": True, "source_ref": source_ref}

        workbook.close()
        return {"deleted": False, "source_ref": source_ref}

